"""YES24 IT/모바일 베스트셀러 데이터를 바탕으로 Excel 대시보드를 생성하는 모듈.

이 모듈은 data 폴더 내의 CSV 데이터를 가공하여 Data 시트에 적재하고,
이를 참조하는 수식(KPI, 출판사 요약, 연도별 집계)과 openpyxl 내장 차트를 포함하는
Dashboard 시트를 생성하여 시각적으로 정돈된 프리미엄 엑셀 대시보드를 구축합니다.
"""

import os
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 파일 경로 정의
SRC_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SRC_DIR)
CSV_FILE = os.path.join(PROJECT_DIR, "data", "yes24_it_mobile_bestsellers.csv")
OUTPUT_EXCEL = os.path.join(PROJECT_DIR, "data", "yes24_bestsellers_dashboard.xlsx")


def preprocess_data(filepath: str) -> pd.DataFrame:
    """CSV 데이터를 로드하여 엑셀 형식에 맞게 자료형을 전처리합니다.

    Args:
        filepath (str): 원본 CSV 파일 경로

    Returns:
        pd.DataFrame: 전처리 완료된 데이터프레임
    """
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"데이터 파일이 없습니다: {filepath}")

    df = pd.read_csv(filepath)

    # 1. 가격 정보 전처리 (콤마 제거 후 숫자 변환)
    for col in ["판매가", "정가"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(",", "")
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).astype(int)

    # 2. 할인율 전처리 (정수형 백분율로 저장하기 위해 100으로 나눔)
    if "할인율" in df.columns:
        df["할인율_값"] = df["할인율"].astype(str).str.replace("%", "")
        df["할인율_값"] = pd.to_numeric(df["할인율_값"], errors="coerce").fillna(0).astype(float)
        # 엑셀 백분율 스타일 적용을 위해 0.0 ~ 1.0 범위로 변환
        df["할인율_값"] = df["할인율_값"] / 100.0
    else:
        df["할인율_값"] = 0.0

    # 3. 판매지수 전처리 (콤마 제거 후 숫자 변환)
    if "판매지수" in df.columns:
        df["판매지수_값"] = df["판매지수"].astype(str).str.replace(",", "")
        df["판매지수_값"] = pd.to_numeric(df["판매지수_값"], errors="coerce").fillna(0).astype(int)
    else:
        df["판매지수_값"] = 0

    # 4. 출판일에서 연도 추출 (예: '2026년 05월' -> 2026)
    date_col = "출판일" if "출판일" in df.columns else ("출간일" if "출간일" in df.columns else None)
    if date_col:
        df["출판연도"] = df[date_col].astype(str).apply(
            lambda x: int(re.search(r"(\d{4})년", x).group(1)) if re.search(r"(\d{4})년", x) else "기타"
        )
    else:
        df["출판연도"] = "기타"

    # 5. 컬럼명 정돈 및 기본 대체값 설정
    if "도서명" not in df.columns and "제목" in df.columns:
        df["도서명"] = df["제목"]
    df["내용"] = df["내용"].fillna("소개 정보 없음") if "내용" in df.columns else "소개 정보 없음"
    df["이미지"] = df["이미지"].fillna("") if "이미지" in df.columns else ""

    return df


def create_excel_dashboard():
    """전처리된 데이터를 바탕으로 엑셀 대시보드를 생성합니다.

    두 개의 시트(Dashboard, Data)를 생성하고 스타일링, 집계 수식, 차트를 적용합니다.
    """
    df = preprocess_data(CSV_FILE)
    wb = Workbook()

    # -------------------------------------------------------------
    # 1. Data 시트 구성 (Raw Data 적재)
    # -------------------------------------------------------------
    ws_data = wb.create_sheet(title="Data")
    
    # 헤더 정의
    headers = [
        "순위", "도서번호", "도서명", "링크", "저자", 
        "출판사", "출판일", "판매가", "정가", "할인율", 
        "판매지수", "출판연도", "내용", "이미지"
    ]
    ws_data.append(headers)

    # 데이터 프레임 로우 쓰기
    for _, row in df.iterrows():
        ws_data.append([
            int(row.get("순위", 0)),
            str(row.get("도서번호", "")),
            str(row.get("도서명", "")),
            str(row.get("링크", "")),
            str(row.get("저자", "")),
            str(row.get("출판사", "")),
            str(row.get("출판일", row.get("출간일", ""))),
            int(row.get("판매가", 0)),
            int(row.get("정가", 0)),
            float(row.get("할인율_값", 0.0)),
            int(row.get("판매지수_값", 0)),
            row.get("출판연도", "기타"),
            str(row.get("내용", "")),
            str(row.get("이미지", ""))
        ])

    max_data_row = len(df) + 1  # 헤더 포함 마지막 행 위치

    # 스타일 설정을 위한 도구 정의
    font_main = Font(name="맑은 고딕", size=10)
    font_bold = Font(name="맑은 고딕", size=10, bold=True)
    font_header = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
    
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Navy
    fill_zebra = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid") # 옅은 회청색
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # Data 시트 스타일 적용
    for r_idx in range(1, max_data_row + 1):
        for c_idx in range(1, len(headers) + 1):
            cell = ws_data.cell(row=r_idx, column=c_idx)
            cell.font = font_main
            cell.border = border_all
            
            # 헤더 행 디자인
            if r_idx == 1:
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = align_center
            else:
                # 홀수 행에만 지브라 패턴(스트라이프 효과) 적용
                if r_idx % 2 == 1:
                    cell.fill = fill_zebra
                
                # 열별 정렬 및 숫자 표시 형식 지정
                if c_idx in [1, 2, 7, 12]:  # 순위, 도서번호, 출판일, 출판연도
                    cell.alignment = align_center
                elif c_idx in [8, 9]:  # 판매가, 정가
                    cell.number_format = "\\#,##0"
                    cell.alignment = align_right
                elif c_idx == 10:  # 할인율
                    cell.number_format = "0.0%"
                    cell.alignment = align_right
                elif c_idx == 11:  # 판매지수
                    cell.number_format = "#,##0"
                    cell.alignment = align_right
                else:
                    cell.alignment = align_left

    # 열 너비 자동 최적화 (너무 긴 '내용', '이미지', '링크' 컬럼은 적절한 크기로 제한)
    for col in ws_data.columns:
        col_letter = get_column_letter(col[0].column)
        if col[0].column in [3, 13, 14]:  # 도서명, 내용, 이미지 링크는 고정 너비 지정
            ws_data.column_dimensions[col_letter].width = 30
        elif col[0].column == 4:  # 링크 컬럼
            ws_data.column_dimensions[col_letter].width = 15
        else:
            max_len = max(len(str(cell.value or '')) for cell in col)
            ws_data.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # -------------------------------------------------------------
    # 2. Dashboard 시트 구성 (대시보드 메인)
    # -------------------------------------------------------------
    ws_dash = wb.active # 첫 번째 기본 시트를 대시보드로 활용
    ws_dash.title = "Dashboard"
    ws_dash.views.sheetView[0].showGridLines = True  # 그리드 라인 표시 활성화

    # 2-1. 대시보드 메인 타이틀 생성 및 스타일링
    ws_dash.merge_cells("A1:F2")
    title_cell = ws_dash["A1"]
    title_cell.value = "YES24 IT/모바일 베스트셀러 분석 대시보드"
    title_cell.font = Font(name="맑은 고딕", size=16, bold=True, color="1F497D")
    title_cell.fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid") # 매우 연한 파란색
    title_cell.alignment = align_center
    
    # 타이틀 영역 테두리 설정
    thick_border_bottom = Border(bottom=Side(border_style="medium", color="1F497D"))
    for row in ws_dash["A1:F2"]:
        for cell in row:
            cell.border = thick_border_bottom

    # 2-2. 주요 지표 (KPI) 영역 구성 (B4:E6)
    kpis = [
        {"title": "총 베스트셀러 도서 수", "formula": f"=COUNTA(Data!C2:C{max_data_row})", "format": "#,##0\"권\"", "col_idx": 2},
        {"title": "평균 판매가", "formula": f"=AVERAGE(Data!H2:H{max_data_row})", "format": "\\#,##0", "col_idx": 3},
        {"title": "평균 할인율", "formula": f"=AVERAGE(Data!J2:J{max_data_row})", "format": "0.0%", "col_idx": 4},
        {"title": "최고 판매지수", "formula": f"=MAX(Data!K2:K{max_data_row})", "format": "#,##0", "col_idx": 5}
    ]

    fill_kpi_lbl = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    fill_kpi_val = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid")
    border_kpi = Border(
        left=Side(border_style="thin", color="A6B9D0"),
        right=Side(border_style="thin", color="A6B9D0"),
        top=Side(border_style="thin", color="A6B9D0"),
        bottom=Side(border_style="thin", color="A6B9D0")
    )

    for kpi in kpis:
        c_letter = get_column_letter(kpi["col_idx"])
        
        # 1) KPI 타이틀 셀 설정
        lbl_cell = ws_dash[f"{c_letter}4"]
        lbl_cell.value = kpi["title"]
        lbl_cell.font = Font(name="맑은 고딕", size=9, bold=True, color="595959")
        lbl_cell.alignment = align_center
        lbl_cell.fill = fill_kpi_lbl
        lbl_cell.border = border_kpi

        # 2) KPI 값 수식 셀 설정
        val_cell = ws_dash[f"{c_letter}5"]
        val_cell.value = kpi["formula"]
        val_cell.font = Font(name="맑은 고딕", size=14, bold=True, color="1F497D")
        val_cell.alignment = align_center
        val_cell.fill = fill_kpi_val
        val_cell.border = border_kpi
        val_cell.number_format = kpi["format"]

    # 2-3. 출판사 점유율 요약 테이블 (Top 10)
    ws_dash["A8"] = "🏢 출판사별 베스트셀러 요약 (Top 10)"
    ws_dash["A8"].font = Font(name="맑은 고딕", size=11, bold=True, color="1F497D")
    
    table_headers_pub = ["출판사", "도서 수", "평균 판매지수"]
    for idx, th in enumerate(table_headers_pub):
        cell = ws_dash.cell(row=9, column=idx + 1, value=th)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all

    # 빈도가 가장 높은 상위 10개 출판사 추출
    top_publishers = df["출판사"].value_counts().head(10).index.tolist()
    # 10개가 안 채워질 경우 빈 문자열로 보완
    while len(top_publishers) < 10:
        top_publishers.append("")

    for i, pub in enumerate(top_publishers):
        r = 10 + i
        # A열: 출판사명
        ws_dash.cell(row=r, column=1, value=pub).alignment = align_left
        # B열: 도서 수 공식 (=COUNTIF)
        cell_count = ws_dash.cell(row=r, column=2, value=f"=COUNTIF(Data!$F$2:$F${max_data_row}, A{r})")
        cell_count.alignment = align_right
        cell_count.number_format = "#,##0"
        # C열: 평균 판매지수 공식 (=AVERAGEIF)
        cell_avg = ws_dash.cell(row=r, column=3, value=f"=AVERAGEIF(Data!$F$2:$F${max_data_row}, A{r}, Data!$K$2:$K${max_data_row})")
        cell_avg.alignment = align_right
        cell_avg.number_format = "#,##0"

        # 행 스타일 일괄 지정
        for col_idx in range(1, 4):
            c = ws_dash.cell(row=r, column=col_idx)
            c.font = font_main
            c.border = border_all
            if r % 2 == 1:
                c.fill = fill_zebra

    # 2-4. 연도별 도서 등록 추이 테이블
    ws_dash["E8"] = "📅 연도별 도서 등록 분포"
    ws_dash["E8"].font = Font(name="맑은 고딕", size=11, bold=True, color="1F497D")

    table_headers_yr = ["출판연도", "도서 수"]
    for idx, th in enumerate(table_headers_yr):
        cell = ws_dash.cell(row=9, column=idx + 5, value=th)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_all

    # 데이터에 존재하는 고유 출판연도 정렬 및 추출 (기타 제외)
    unique_years = [y for y in df["출판연도"].unique() if y != "기타" and pd.notna(y)]
    unique_years = sorted(list(set(unique_years)))
    
    # 최근 6개년 위주로 구성 (테이블 균형 확보)
    if len(unique_years) > 6:
        unique_years = unique_years[-6:]

    for i, yr in enumerate(unique_years):
        r = 10 + i
        # E열: 연도
        ws_dash.cell(row=r, column=5, value=yr).alignment = align_center
        # F열: 도서 수 공식 (=COUNTIF)
        cell_yr_cnt = ws_dash.cell(row=r, column=6, value=f"=COUNTIF(Data!$L$2:$L${max_data_row}, E{r})")
        cell_yr_cnt.alignment = align_right
        cell_yr_cnt.number_format = "#,##0"

        # 행 스타일 일괄 지정
        for col_idx in range(5, 7):
            c = ws_dash.cell(row=r, column=col_idx)
            c.font = font_main
            c.border = border_all
            if r % 2 == 1:
                c.fill = fill_zebra

    # 2-5. openpyxl 차트 생성 및 삽입
    
    # 1) 출판사별 도서 수 2D 세로 막대 그래프 (BarChart)
    chart_bar = BarChart()
    chart_bar.type = "col"
    chart_bar.style = 10
    chart_bar.title = "출판사별 베스트셀러 도서 수 (Top 10)"
    chart_bar.y_axis.title = "도서 수 (권)"
    chart_bar.x_axis.title = "출판사"
    chart_bar.height = 10
    chart_bar.width = 15

    # 데이터 범위 지정 (도서 수 열: B9:B19)
    data_bar = Reference(ws_dash, min_col=2, min_row=9, max_row=19)
    # 카테고리 범위 지정 (출판사 열: A10:A19)
    cats_bar = Reference(ws_dash, min_col=1, min_row=10, max_row=19)

    chart_bar.add_data(data_bar, titles_from_data=True)
    chart_bar.set_categories(cats_bar)
    chart_bar.legend = None # 단일 계열이므로 범례 생략

    # 2) 연도별 도서 수 꺾은선 그래프 (LineChart)
    chart_line = LineChart()
    chart_line.title = "연도별 베스트셀러 등록 추이"
    chart_line.style = 13
    chart_line.y_axis.title = "도서 수 (권)"
    chart_line.x_axis.title = "연도"
    chart_line.height = 10
    chart_line.width = 13

    # 데이터 범위 지정 (도서 수 열: F9:F9+len(unique_years))
    data_line = Reference(ws_dash, min_col=6, min_row=9, max_row=9 + len(unique_years))
    # 카테고리 범위 지정 (연도 열: E10:E9+len(unique_years))
    cats_line = Reference(ws_dash, min_col=5, min_row=10, max_row=9 + len(unique_years))

    chart_line.add_data(data_line, titles_from_data=True)
    chart_line.set_categories(cats_line)
    chart_line.legend = None

    # 차트 배치
    ws_dash.add_chart(chart_bar, "A21")
    ws_dash.add_chart(chart_line, "E21")

    # 대시보드 시트 컬럼 너비 조정
    ws_dash.column_dimensions["A"].width = 25
    ws_dash.column_dimensions["B"].width = 15
    ws_dash.column_dimensions["C"].width = 18
    ws_dash.column_dimensions["D"].width = 5
    ws_dash.column_dimensions["E"].width = 18
    ws_dash.column_dimensions["F"].width = 15

    # 최종 엑셀 파일 저장
    wb.save(OUTPUT_EXCEL)
    print(f"[SUCCESS] 엑셀 대시보드가 정상적으로 생성되었습니다: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    create_excel_dashboard()
